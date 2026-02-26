"""
Kira Senaryo — TCMB Tüketici Fiyatları verisiyle TÜFE bazlı kira hesaplama
Mantık:
  - Her yıl YALNIZCA sözleşme yıldönümünde artış uygulanır
  - Yıldönümü henüz gelmediyse: mevcut kira devam eder (artış YOK)
  - Yıldönümüne ≤1 ay kaldıysa: bir önceki ayın TÜFE'si ile "erken bildirim tahmini" gösterilir
  - Excel çıktısı: firma adı + tüm tablo
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date, datetime
import threading, os
from pathlib import Path

AYLAR_TR = ["Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
            "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

# ── TCMB: Her ayın 12 aylık yıllık TÜFE oranı ───────────────────────────────
AYLIK_YILLIK = {
    (2005,1):9.24,(2005,2):8.69,(2005,3):7.94,(2005,4):8.18,(2005,5):8.70,
    (2005,6):8.95,(2005,7):7.82,(2005,8):7.91,(2005,9):7.99,(2005,10):7.52,
    (2005,11):7.61,(2005,12):7.72,
    (2006,1):7.93,(2006,2):8.15,(2006,3):8.16,(2006,4):8.83,(2006,5):9.86,
    (2006,6):10.12,(2006,7):11.69,(2006,8):10.26,(2006,9):10.55,(2006,10):9.98,
    (2006,11):9.86,(2006,12):9.65,
    (2007,1):9.93,(2007,2):10.16,(2007,3):10.86,(2007,4):10.72,(2007,5):9.23,
    (2007,6):8.60,(2007,7):6.90,(2007,8):7.39,(2007,9):7.12,(2007,10):7.70,
    (2007,11):8.40,(2007,12):8.39,
    (2008,1):8.17,(2008,2):9.10,(2008,3):9.15,(2008,4):9.66,(2008,5):10.74,
    (2008,6):10.61,(2008,7):12.06,(2008,8):11.77,(2008,9):11.13,(2008,10):11.99,
    (2008,11):10.76,(2008,12):10.06,
    (2009,1):9.50,(2009,2):7.73,(2009,3):7.89,(2009,4):6.13,(2009,5):5.24,
    (2009,6):5.73,(2009,7):5.39,(2009,8):5.33,(2009,9):5.27,(2009,10):5.08,
    (2009,11):5.53,(2009,12):6.53,
    (2010,1):8.19,(2010,2):10.13,(2010,3):9.56,(2010,4):10.19,(2010,5):9.10,
    (2010,6):8.37,(2010,7):7.58,(2010,8):8.33,(2010,9):9.24,(2010,10):8.62,
    (2010,11):7.29,(2010,12):6.40,
    (2011,1):4.90,(2011,2):4.16,(2011,3):3.99,(2011,4):4.26,(2011,5):7.17,
    (2011,6):6.24,(2011,7):6.31,(2011,8):6.65,(2011,9):6.15,(2011,10):7.66,
    (2011,11):9.48,(2011,12):10.45,
    (2012,1):10.61,(2012,2):10.43,(2012,3):10.43,(2012,4):11.14,(2012,5):8.28,
    (2012,6):8.87,(2012,7):9.07,(2012,8):8.88,(2012,9):9.19,(2012,10):7.80,
    (2012,11):6.37,(2012,12):6.16,
    (2013,1):7.31,(2013,2):7.03,(2013,3):7.29,(2013,4):6.13,(2013,5):6.51,
    (2013,6):8.30,(2013,7):8.88,(2013,8):8.17,(2013,9):7.88,(2013,10):7.71,
    (2013,11):7.32,(2013,12):7.40,
    (2014,1):7.75,(2014,2):7.89,(2014,3):8.39,(2014,4):9.38,(2014,5):9.66,
    (2014,6):9.16,(2014,7):9.32,(2014,8):9.54,(2014,9):8.86,(2014,10):8.96,
    (2014,11):9.15,(2014,12):8.17,
    (2015,1):7.24,(2015,2):7.55,(2015,3):7.61,(2015,4):7.91,(2015,5):8.09,
    (2015,6):7.20,(2015,7):6.81,(2015,8):7.14,(2015,9):7.95,(2015,10):7.58,
    (2015,11):8.10,(2015,12):8.81,
    (2016,1):9.58,(2016,2):8.78,(2016,3):7.46,(2016,4):6.57,(2016,5):6.58,
    (2016,6):7.64,(2016,7):8.79,(2016,8):8.05,(2016,9):7.28,(2016,10):7.16,
    (2016,11):7.00,(2016,12):8.53,
    (2017,1):9.22,(2017,2):10.13,(2017,3):11.29,(2017,4):11.87,(2017,5):11.72,
    (2017,6):10.90,(2017,7):9.79,(2017,8):10.68,(2017,9):11.20,(2017,10):11.90,
    (2017,11):12.98,(2017,12):11.92,
    (2018,1):10.35,(2018,2):10.26,(2018,3):10.23,(2018,4):10.85,(2018,5):12.15,
    (2018,6):15.39,(2018,7):15.85,(2018,8):17.90,(2018,9):24.52,(2018,10):25.24,
    (2018,11):21.62,(2018,12):20.30,
    (2019,1):20.35,(2019,2):19.67,(2019,3):19.71,(2019,4):19.50,(2019,5):18.71,
    (2019,6):15.72,(2019,7):16.65,(2019,8):15.01,(2019,9):9.26,(2019,10):8.55,
    (2019,11):10.56,(2019,12):11.84,
    (2020,1):12.15,(2020,2):12.37,(2020,3):11.86,(2020,4):10.94,(2020,5):11.39,
    (2020,6):12.62,(2020,7):11.76,(2020,8):11.77,(2020,9):11.75,(2020,10):11.89,
    (2020,11):14.03,(2020,12):14.60,
    (2021,1):14.97,(2021,2):15.61,(2021,3):16.19,(2021,4):17.14,(2021,5):16.59,
    (2021,6):17.53,(2021,7):18.95,(2021,8):19.25,(2021,9):19.58,(2021,10):19.89,
    (2021,11):21.31,(2021,12):36.08,
    (2022,1):48.69,(2022,2):54.44,(2022,3):61.14,(2022,4):69.97,(2022,5):73.50,
    (2022,6):78.62,(2022,7):79.60,(2022,8):80.21,(2022,9):83.45,(2022,10):85.51,
    (2022,11):84.39,(2022,12):64.27,
    (2023,1):57.68,(2023,2):55.18,(2023,3):50.51,(2023,4):43.68,(2023,5):39.59,
    (2023,6):38.21,(2023,7):47.83,(2023,8):58.94,(2023,9):61.53,(2023,10):61.36,
    (2023,11):61.98,(2023,12):64.77,
    (2024,1):64.86,(2024,2):67.07,(2024,3):68.50,(2024,4):69.80,(2024,5):75.45,
    (2024,6):71.60,(2024,7):61.78,(2024,8):51.97,(2024,9):49.38,(2024,10):48.58,
    (2024,11):47.09,(2024,12):44.38,
    (2025,1):42.12,(2025,2):39.05,(2025,3):38.10,(2025,4):37.86,(2025,5):35.41,
    (2025,6):35.05,(2025,7):33.52,(2025,8):32.95,(2025,9):33.29,(2025,10):32.87,
    (2025,11):31.07,(2025,12):30.89,
    (2026,1):30.65,  # En son TCMB verisi (Ocak 2026)
}


def _onceki_ay(yil, ay):
    """Bir önceki ayı döndür."""
    if ay == 1:
        return yil - 1, 12
    return yil, ay - 1


def _tufe_al(yil, ay):
    """Belirtilen yıl/ay için TCMB yıllık TÜFE oranını döndür."""
    return AYLIK_YILLIK.get((yil, ay), 0.0)


def hesapla(baslangic: float, bas_gun: int, bas_ay: int, bas_yil: int) -> dict:
    """
    Kira artış hesabı:
    - Yıldönümü = her yıl başlangıç ayının aynı günü
    - Yıldönümü geçtiyse: o ayın TÜFE'si uygulanır, tutar güncellenir
    - Yıldönümü henüz gelmediyse: artış YOK, tutar aynı kalır
    - Yıldönümüne ≤1 ay kaldıysa: erken bildirim tahmini gösterilir
    """
    bugun = date.today()
    try:
        bas_gun_gercek = min(bas_gun, 28)
        bas_tarih = date(bas_yil, bas_ay, bas_gun_gercek)
    except:
        return {"hata": "Geçersiz tarih!"}

    if bas_tarih >= bugun:
        return {"hata": "Başlangıç tarihi bugünden önce olmalı!"}

    satirlar = []
    tutar = baslangic
    erken_bildirim = None  # Varsa doldurulacak

    # Başlangıç satırı
    satirlar.append({
        "yil": bas_yil, "tarih": bas_tarih.strftime("%d.%m.%Y"),
        "tutar": round(tutar, 2), "tufe": 0.0, "artis": 0.0,
        "aciklama": "Sözleşme Başlangıcı", "tip": "bas",
    })

    for yil in range(bas_yil + 1, bugun.year + 4):
        gun_gercek = min(bas_gun, 28)
        try:
            yildonumu = date(yil, bas_ay, gun_gercek)
        except:
            yildonumu = date(yil, bas_ay, 28)

        if yildonumu <= bugun:
            # Yıldönümü geçti → artış uygulandı
            tufe_oran = _tufe_al(yil, bas_ay)
            # Eğer o ay için veri yoksa bir önceki ayı dene
            if tufe_oran == 0.0:
                py, pm = _onceki_ay(yil, bas_ay)
                tufe_oran = _tufe_al(py, pm)
            yeni_tutar = tutar * (1 + tufe_oran / 100)
            satirlar.append({
                "yil": yil, "tarih": yildonumu.strftime("%d.%m.%Y"),
                "tutar": round(yeni_tutar, 2), "tufe": tufe_oran,
                "artis": round(yeni_tutar - tutar, 2),
                "aciklama": f"TCMB: {AYLAR_TR[bas_ay-1]} {yil} yıllık TÜFE",
                "tip": "gercek",
            })
            tutar = yeni_tutar

        else:
            # Yıldönümü henüz gelmedi
            # Artışa kaç gün kaldı?
            kalan_gun = (yildonumu - bugun).days

            if kalan_gun <= 35:
                # ≤35 gün kaldı → erken bildirim:
                # En son AÇIKLANMIŞ TÜFE oranını bul (bugüne kadar)
                # Önce: yıldönümü ayının TÜFE'si açıklandı mı?
                tufe_yon_ay = _tufe_al(yil, bas_ay)
                if tufe_yon_ay > 0.0:
                    # Yıldönümü ayının TÜFE'si zaten var → bunu kullan
                    tufe_onceki = tufe_yon_ay
                    tufe_ay_adi = f"{AYLAR_TR[bas_ay-1]} {yil}"
                else:
                    # Henüz açıklanmadı → geriye giderek son açıklanan TÜFE'yi bul
                    bulunan_yil, bulunan_ay = yil, bas_ay
                    tufe_onceki = 0.0
                    tufe_ay_adi = "?"
                    for _ in range(24):  # max 24 ay geri git
                        bulunan_yil, bulunan_ay = _onceki_ay(bulunan_yil, bulunan_ay)
                        v = _tufe_al(bulunan_yil, bulunan_ay)
                        if v > 0.0:
                            tufe_onceki = v
                            tufe_ay_adi = f"{AYLAR_TR[bulunan_ay-1]} {bulunan_yil}"
                            break
                if tufe_onceki == 0.0:
                    # TÜFE verisi yoksa erken bildirim gösterme
                    satirlar.append({
                        "yil": yil, "tarih": yildonumu.strftime("%d.%m.%Y") + " ⟳",
                        "tutar": round(tutar, 2), "tufe": 0.0, "artis": 0.0,
                        "aciklama": f"Erken bildirim ({kalan_gun} gün kaldı) — TÜFE verisi bekleniyor",
                        "tip": "erken",
                    })
                else:
                    tahmini_tutar = tutar * (1 + tufe_onceki / 100)
                    erken_bildirim = {
                        "tarih": yildonumu.strftime("%d.%m.%Y"),
                        "kalan_gun": kalan_gun,
                        "tufe_ay": tufe_ay_adi,
                        "tufe_oran": tufe_onceki,
                        "mevcut_tutar": round(tutar, 2),
                        "tahmini_tutar": round(tahmini_tutar, 2),
                        "tahmini_artis": round(tahmini_tutar - tutar, 2),
                    }
                    satirlar.append({
                        "yil": yil, "tarih": yildonumu.strftime("%d.%m.%Y") + " ⟳",
                        "tutar": round(tahmini_tutar, 2), "tufe": tufe_onceki,
                        "artis": round(tahmini_tutar - tutar, 2),
                        "aciklama": f"Erken bildirim tahmini ({kalan_gun} gün kaldı) — {tufe_ay_adi} TÜFE",
                        "tip": "erken",
                    })
            else:
                # Uzak gelecek → beklenen (mevcut tutar devam)
                satirlar.append({
                    "yil": yil, "tarih": yildonumu.strftime("%d.%m.%Y") + " ⟳",
                    "tutar": round(tutar, 2), "tufe": 0.0, "artis": 0.0,
                    "aciklama": "Beklenen (henüz uygulanmadı)",
                    "tip": "beklenen",
                })

            # 2 yıldan fazla ileriye bakma
            if yil > bugun.year + 1:
                break

    gecerli = tutar
    return {
        "hata": "", "baslangic": baslangic, "gecerli": round(gecerli, 2),
        "artis_tl": round(gecerli - baslangic, 2),
        "artis_pct": round((gecerli / baslangic - 1) * 100, 1) if baslangic else 0,
        "bas_tarih": bas_tarih.strftime("%d.%m.%Y"),
        "satirlar": satirlar, "erken_bildirim": erken_bildirim,
    }


# ── Excel çıktısı ────────────────────────────────────────────────────────────
def excel_olustur(sonuc: dict, musteri_ad: str, kayit_yolu: str):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Kira Artış Senaryosu"

    # Renkler
    C_HEADER  = "1E3A5F"   # koyu lacivert
    C_BAS     = "2D4A6B"   # başlangıç satırı
    C_GERCEK  = "1A3A2A"   # gerçekleşen — koyu yeşil
    C_SIMDI   = "0D5C2E"   # bugün geçerli
    C_ERKEN   = "4A3B00"   # erken bildirim — koyu sarı
    C_BEKLENEN= "2A2A3A"   # beklenen — koyu mor
    C_ALT     = "1A2535"   # alternatif satır
    C_NORMAL  = "0F1E2D"   # normal satır arka plan

    def stil(cell, bold=False, fg="FFFFFF", bg=None, align="center", size=10):
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg, end_color=bg)

    def border_all(cell, renk="2C4F7C"):
        s = Side(style="thin", color=renk)
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    # ── Başlık ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "KİRA ARTIŞ SENARYOSU — TCMB TÜFE BAZLI"
    stil(ws["A1"], bold=True, fg="FFFFFF", bg="0F2D4A", size=14)
    ws.row_dimensions[1].height = 28

    # Firma + tarih bilgisi
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Firma / Müşteri: {musteri_ad or '—'}"
    stil(ws["A2"], bold=True, fg="4FC3F7", bg="0A1929", size=11, align="left")

    ws.merge_cells("E2:G2")
    ws["E2"] = f"Rapor Tarihi: {date.today().strftime('%d.%m.%Y')}  |  Kaynak: TCMB Tüketici Fiyatları"
    stil(ws["E2"], fg="888888", bg="0A1929", size=9, align="right")
    ws.row_dimensions[2].height = 22

    # Özet kartları
    ws.merge_cells("A3:B3")
    ws["A3"] = "Başlangıç Kirası"
    stil(ws["A3"], bold=True, fg="AAAAAA", bg="12243A", size=9)
    ws.merge_cells("A4:B4")
    ws["A4"] = f"{sonuc['baslangic']:,.2f} TL  ({sonuc['bas_tarih']})"
    stil(ws["A4"], bold=True, fg="E0F7FA", bg="12243A", size=12)

    ws.merge_cells("C3:D3")
    ws["C3"] = "Bugün Geçerli Kira"
    stil(ws["C3"], bold=True, fg="AAAAAA", bg="0D2E1A", size=9)
    ws.merge_cells("C4:D4")
    ws["C4"] = f"{sonuc['gecerli']:,.2f} TL"
    stil(ws["C4"], bold=True, fg="69F0AE", bg="0D2E1A", size=14)

    ws.merge_cells("E3:E4")
    ws["E3"] = f"+{sonuc['artis_tl']:,.2f} TL"
    stil(ws["E3"], bold=True, fg="FF8A65", bg="2A1A0A", size=12)

    ws.merge_cells("F3:G4")
    ws["F3"] = f"%{sonuc['artis_pct']:,.1f} toplam artış"
    stil(ws["F3"], bold=True, fg="FF8A65", bg="2A1A0A", size=12)

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 26

    # Erken bildirim kutusu
    eb = sonuc.get("erken_bildirim")
    if eb:
        ws.merge_cells("A5:G5")
        ws["A5"] = (f"⚠  ERKen BİLDİRİM: Kira artışına {eb['kalan_gun']} gün kaldı. "
                    f"{eb['tufe_ay']} TÜFE %{eb['tufe_oran']:.2f} uygulandığında → "
                    f"Tahmini yeni kira: {eb['tahmini_tutar']:,.2f} TL  "
                    f"(+{eb['tahmini_artis']:,.2f} TL artış)")
        stil(ws["A5"], bold=True, fg="FFF176", bg="3A2E00", size=10, align="left")
        ws.row_dimensions[5].height = 22
        tablo_bas = 7
    else:
        ws.row_dimensions[5].height = 0
        tablo_bas = 6

    # Boş ayıraç
    ws.row_dimensions[tablo_bas - 1].height = 6

    # Tablo başlıkları
    basliklar = ["Yıldönümü Tarihi", "Yıl", "TÜFE (Yıllık %)", "Kira (TL)",
                 "Artış (TL)", "Artış (%)", "Açıklama"]
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=tablo_bas, column=col, value=baslik)
        stil(cell, bold=True, fg="FFFFFF", bg=C_HEADER, size=10)
        border_all(cell, "1E5A8E")
    ws.row_dimensions[tablo_bas].height = 20

    # Tablo verileri
    son_gercek_row = tablo_bas
    for i, s in enumerate(sonuc["satirlar"]):
        row = tablo_bas + 1 + i
        tip = s["tip"]

        bg = {"bas": C_BAS, "gercek": C_GERCEK, "erken": C_ERKEN,
              "beklenen": C_BEKLENEN}.get(tip, C_NORMAL)
        if tip == "gercek":
            # En son gerçek satırı vurgula
            son_gercek_row = row

        tufe_str = f"%{s['tufe']:.2f}" if s["tufe"] > 0 else "—"
        artis_str = f"+{s['artis']:,.2f}" if s["artis"] > 0 else "—"
        artis_pct = ""
        if s["artis"] > 0 and s["tutar"] > 0:
            baz = s["tutar"] - s["artis"]
            artis_pct = f"%{(s['artis']/baz*100):.2f}" if baz > 0 else ""

        degerler = [s["tarih"], str(s["yil"]), tufe_str,
                    f"{s['tutar']:,.2f}", artis_str, artis_pct, s["aciklama"]]

        for col, val in enumerate(degerler, 1):
            cell = ws.cell(row=row, column=col, value=val)
            fg_renk = {
                "bas": "90A4AE", "gercek": "A5D6A7",
                "erken": "FFF176", "beklenen": "9FA8DA"
            }.get(tip, "E0F7FA")
            if tip == "gercek" and row == son_gercek_row:
                fg_renk = "69F0AE"
                cell.font = Font(name="Calibri", bold=True, color=fg_renk, size=10)
            else:
                cell.font = Font(name="Calibri",
                                  bold=(tip == "erken"),
                                  color=fg_renk,
                                  italic=(tip == "beklenen"),
                                  size=10)
            cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            border_all(cell, "1E3A5F")
        ws.row_dimensions[row].height = 18

    # En son geçerli satırı koyu yeşil vurgula
    if son_gercek_row > tablo_bas:
        for col in range(1, 8):
            cell = ws.cell(row=son_gercek_row, column=col)
            cell.fill = PatternFill("solid", start_color=C_SIMDI, end_color=C_SIMDI)

    # Kolon genişlikleri
    genislikler = [18, 8, 15, 14, 14, 12, 45]
    for col, w in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Alt not
    son_row = tablo_bas + 1 + len(sonuc["satirlar"]) + 1
    ws.merge_cells(f"A{son_row}:G{son_row}")
    ws[f"A{son_row}"] = ("Kaynak: TCMB Tüketici Fiyatları — tcmb.gov.tr  |  "
                          "Her yıl sözleşme yıldönümünde (aynı gün ve ay) artış uygulanır.  |  "
                          "⟳ = Yıldönümü henüz gelmedi, tahmini değerdir.")
    stil(ws[f"A{son_row}"], fg="666666", bg="080F18", size=8, align="left")

    wb.save(kayit_yolu)


# ── Widget ───────────────────────────────────────────────────────────────────
class KiraSenaryoFrame(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding=6)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)
        self._sonuc = None
        self._build()

    def _build(self):
        # Başlık
        ust = ttk.Frame(self); ust.grid(row=0, column=0, sticky="ew", pady=(0,6))
        ttk.Label(ust, text="TÜFE Bazlı Kira Artış Senaryosu",
                  font=("Segoe UI",12,"bold")).pack(side=tk.LEFT)
        self.lbl_kaynak = ttk.Label(ust, text="  Kaynak: TCMB Tüketici Fiyatları",
                                     font=("Segoe UI",8), foreground="#4fc3f7")
        self.lbl_kaynak.pack(side=tk.LEFT)
        self.lbl_durum = ttk.Label(ust, text="", font=("Segoe UI",8), foreground="#888")
        self.lbl_durum.pack(side=tk.LEFT, padx=6)
        ttk.Button(ust, text="📊 Excel Çıktısı",
                   command=self._excel_al).pack(side=tk.RIGHT, padx=4)
        ttk.Button(ust, text="🌐 TCMB Güncelle",
                   command=self._guncelle).pack(side=tk.RIGHT, padx=4)

        # Giriş
        gf = ttk.LabelFrame(self, text="Kira Bilgileri", padding=10)
        gf.grid(row=1, column=0, sticky="ew", pady=(0,6))

        # Müşteri
        f0 = ttk.Frame(gf); f0.pack(fill=tk.X, pady=2)
        ttk.Label(f0, text="Müşteri / Ad:", width=18, anchor="e").pack(side=tk.LEFT)
        self.v_musteri = tk.StringVar()
        ttk.Entry(f0, textvariable=self.v_musteri, width=22).pack(side=tk.LEFT, padx=6)
        ttk.Label(f0, text="  veya seç:", foreground="#888").pack(side=tk.LEFT)
        self.cb_musteri = ttk.Combobox(f0, width=22, state="readonly")
        self.cb_musteri.pack(side=tk.LEFT, padx=4)
        self.cb_musteri.bind("<<ComboboxSelected>>",
            lambda e: self.v_musteri.set(self.cb_musteri.get().split(" —")[0]))
        self._yukle_musteriler()

        # Kira
        f1 = ttk.Frame(gf); f1.pack(fill=tk.X, pady=2)
        ttk.Label(f1, text="Başlangıç Kirası:", width=18, anchor="e").pack(side=tk.LEFT)
        self.v_tutar = tk.StringVar()
        vc = (self.register(lambda p: p=="" or p.replace(".","",1).isdigit()), "%P")
        ttk.Entry(f1, textvariable=self.v_tutar, width=14, validate="key",
                  validatecommand=vc, font=("Segoe UI",10,"bold")).pack(side=tk.LEFT, padx=6)
        ttk.Label(f1, text="TL", foreground="#4fc3f7", font=("Segoe UI",9,"bold")).pack(side=tk.LEFT)

        # Tarih
        f2 = ttk.Frame(gf); f2.pack(fill=tk.X, pady=2)
        ttk.Label(f2, text="Sözleşme Tarihi:", width=18, anchor="e").pack(side=tk.LEFT)
        bugun = date.today()
        self.v_gun = tk.StringVar(value="1")
        self.v_ay  = tk.StringVar(value=str(bugun.month))
        self.v_yil = tk.StringVar(value=str(bugun.year - 3))
        ttk.Combobox(f2, textvariable=self.v_gun,
                     values=[str(i) for i in range(1,32)],
                     width=4, state="readonly").pack(side=tk.LEFT, padx=(6,1))
        ttk.Label(f2, text="/", foreground="#888").pack(side=tk.LEFT)
        ttk.Combobox(f2, textvariable=self.v_ay,
                     values=[str(i) for i in range(1,13)],
                     width=3, state="readonly").pack(side=tk.LEFT, padx=1)
        ttk.Label(f2, text="/", foreground="#888").pack(side=tk.LEFT)
        ttk.Combobox(f2, textvariable=self.v_yil,
                     values=[str(y) for y in range(2005, bugun.year+1)],
                     width=6, state="readonly").pack(side=tk.LEFT, padx=(1,8))
        self.lbl_ay_adi = ttk.Label(f2, text=AYLAR_TR[bugun.month-1],
                                     foreground="#4fc3f7", font=("Segoe UI",9,"bold"))
        self.lbl_ay_adi.pack(side=tk.LEFT)
        self.v_ay.trace_add("write", lambda *_: self._ay_guncelle())

        # Butonlar
        bf = ttk.Frame(gf); bf.pack(fill=tk.X, pady=(8,0))
        ttk.Button(bf, text="🧮  Hesapla", command=self._hesapla).pack(
            side=tk.LEFT, ipady=5, padx=(0,8))
        ttk.Button(bf, text="🗑 Temizle", command=self._temizle).pack(side=tk.LEFT)

        # Erken bildirim banner
        self.erken_frame = ttk.Frame(self, style="Card.TFrame")
        self.erken_frame.grid(row=2, column=0, sticky="ew", pady=(0,4))
        self.lbl_erken = ttk.Label(self.erken_frame, text="",
                                    font=("Segoe UI",9,"bold"), foreground="#fff176",
                                    background="#3a2e00", wraplength=900, padding=(8,6))
        self.lbl_erken.pack(fill=tk.X)

        # Sonuç kartları
        kf = ttk.Frame(self); kf.grid(row=3, column=0, sticky="ew", pady=(0,6))

        def kart(renk, baslik):
            lf = ttk.LabelFrame(kf, text=baslik, padding=8)
            lf.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
            lbl = ttk.Label(lf, text="—", font=("Segoe UI",13,"bold"), foreground=renk)
            lbl.pack()
            alt = ttk.Label(lf, text="", font=("Segoe UI",7), foreground="#888")
            alt.pack()
            return lbl, alt

        self.k_bas, self.a_bas = kart("#e0f7fa", "Başlangıç Kirası")
        self.k_sim, self.a_sim = kart("#69f0ae", "Bugün Geçerli Kira")
        self.k_tl,  self.a_tl  = kart("#ff8a65", "Toplam Artış (TL)")
        self.k_pct, self.a_pct = kart("#ff8a65", "Toplam Artış (%)")

        # Tablo
        tf = ttk.LabelFrame(self, text="Yıl Yıl TCMB TÜFE ve Kira Artış Tablosu", padding=6)
        tf.grid(row=4, column=0, sticky="nsew", pady=(0,4))
        self.rowconfigure(4, weight=1)
        tf.columnconfigure(0, weight=1); tf.rowconfigure(0, weight=1)

        cols = ("tarih","yil","tufe","tutar","artis","artis_pct","aciklama")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", height=10)
        hdrs = {"tarih":"Yıldönümü","yil":"Yıl","tufe":"TÜFE (Yıllık %)",
                "tutar":"Kira (TL)","artis":"Artış (TL)","artis_pct":"Artış (%)",
                "aciklama":"Açıklama"}
        wdts = {"tarih":140,"yil":50,"tufe":120,"tutar":130,"artis":110,
                "artis_pct":85,"aciklama":260}
        for c in cols:
            self.tree.heading(c, text=hdrs[c])
            self.tree.column(c, width=wdts[c], anchor="center")
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self.tree.tag_configure("bas",      foreground="#90a4ae")
        self.tree.tag_configure("gercek",   foreground="#a5d6a7")
        self.tree.tag_configure("simdi",    foreground="#69f0ae", font=("Segoe UI",9,"bold"))
        self.tree.tag_configure("erken",    foreground="#fff176", font=("Segoe UI",9,"bold"))
        self.tree.tag_configure("beklenen", foreground="#9fa8da", font=("Segoe UI",9,"italic"))

        ttk.Label(self,
            text="ℹ  Artış yalnızca sözleşme yıldönümünde (aynı gün ve ay) uygulanır. "
                 "Yıldönümüne ≤35 gün kaldıysa erken bildirim tahmini gösterilir. "
                 "⟳ = Henüz gerçekleşmedi.",
            font=("Segoe UI",8), foreground="#888",
            wraplength=850).grid(row=5, column=0, sticky="w", pady=(0,2))

    # ── Yardımcılar ───────────────────────────────────────────────────────────
    def _ay_guncelle(self, *_):
        try: self.lbl_ay_adi.config(text=AYLAR_TR[int(self.v_ay.get())-1])
        except: pass

    def _yukle_musteriler(self):
        try:
            from database import fetch_all
            rows = fetch_all("SELECT id, name FROM customers ORDER BY name")
            self.cb_musteri["values"] = [f"{r['name']} — ID:{r['id']}" for r in rows]
        except:
            self.cb_musteri["values"] = []

    def _hesapla(self):
        try:
            tutar = float(self.v_tutar.get().replace(",","."))
            assert tutar > 0
        except:
            messagebox.showwarning("Hata","Geçerli kira tutarı girin!",parent=self); return
        try:
            gun = int(self.v_gun.get())
            ay  = int(self.v_ay.get())
            yil = int(self.v_yil.get())
            date(yil, ay, min(gun,28))
        except:
            messagebox.showwarning("Hata","Geçerli tarih seçin!",parent=self); return

        sonuc = hesapla(tutar, gun, ay, yil)
        if sonuc.get("hata"):
            messagebox.showwarning("Hata", sonuc["hata"], parent=self); return

        self._sonuc = sonuc

        # Kartlar
        musteri = self.v_musteri.get().strip()
        self.k_bas.config(text=f"{sonuc['baslangic']:,.2f} TL")
        self.a_bas.config(text=f"Tarih: {sonuc['bas_tarih']}" +
                          (f"  [{musteri}]" if musteri else ""))
        self.k_sim.config(text=f"{sonuc['gecerli']:,.2f} TL")
        self.a_sim.config(text="Bugün itibarıyla geçerli tutar")
        self.k_tl.config(text=f"+{sonuc['artis_tl']:,.2f} TL")
        self.k_pct.config(text=f"%{sonuc['artis_pct']:,.1f}")

        # Erken bildirim banner
        eb = sonuc.get("erken_bildirim")
        if eb:
            self.lbl_erken.config(
                text=f"⚠  ERKen BİLDİRİM: Kira artışına {eb['kalan_gun']} gün kaldı!  "
                     f"{eb['tufe_ay']} TÜFE %{eb['tufe_oran']:.2f} uygulandığında → "
                     f"Mevcut: {eb['mevcut_tutar']:,.2f} TL  →  "
                     f"Tahmini yeni kira: {eb['tahmini_tutar']:,.2f} TL  "
                     f"(+{eb['tahmini_artis']:,.2f} TL)"
            )
        else:
            self.lbl_erken.config(text="")

        # Tablo
        self.tree.delete(*self.tree.get_children())
        son_gercek_iid = None
        for s in sonuc["satirlar"]:
            tip = s["tip"]
            artis_str = f"+{s['artis']:,.2f} TL" if s["artis"] > 0 else "—"
            tufe_str  = f"%{s['tufe']:.2f}" if s["tufe"] > 0 else "—"
            artis_pct_str = ""
            if s["artis"] > 0:
                baz = s["tutar"] - s["artis"]
                if baz > 0:
                    artis_pct_str = f"%{(s['artis']/baz*100):.2f}"
            iid = self.tree.insert("", tk.END, tags=(tip,), values=(
                s["tarih"], s["yil"], tufe_str,
                f"{s['tutar']:,.2f} TL", artis_str,
                artis_pct_str, s["aciklama"]))
            if tip == "gercek":
                son_gercek_iid = iid

        if son_gercek_iid:
            self.tree.item(son_gercek_iid, tags=("simdi",))
            self.tree.see(son_gercek_iid)

    def _excel_al(self):
        if not self._sonuc:
            messagebox.showwarning("Uyarı","Önce hesaplama yapın!",parent=self); return
        musteri = self.v_musteri.get().strip() or "Musteri"
        dosya_adi = f"KiraArtis_{musteri.replace(' ','_')}_{date.today().strftime('%Y%m%d')}.xlsx"
        yol = filedialog.asksaveasfilename(
            title="Excel Kaydet", defaultextension=".xlsx",
            initialfile=dosya_adi,
            filetypes=[("Excel","*.xlsx"),("Tüm","*.*")])
        if not yol: return
        try:
            excel_olustur(self._sonuc, musteri, yol)
            self.lbl_durum.config(text=f"✓ Excel kaydedildi: {Path(yol).name}",
                                   foreground="#69f0ae")
            if messagebox.askyesno("Hazır","Excel dosyası oluşturuldu. Şimdi açılsın mı?",
                                    parent=self):
                if os.name == "nt": os.startfile(yol)
                else: __import__("subprocess").Popen(["xdg-open", yol])
        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror("Hata", str(e), parent=self)

    def _guncelle(self):
        self.lbl_durum.config(text="⏳ TCMB'den güncelleniyor...", foreground="#fff176")
        threading.Thread(target=self._guncelle_thread, daemon=True).start()

    def _guncelle_thread(self):
        try:
            import urllib.request, re
            url = ("https://www.tcmb.gov.tr/wps/wcm/connect/TR/TCMB+TR/"
                   "Main+Menu/Istatistikler/Enflasyon+Verileri/Tuketici+Fiyatlari")
            req = urllib.request.Request(url, headers={
                "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120"})
            html = urllib.request.urlopen(req, timeout=10).read().decode("utf-8","ignore")
            pattern = r'\|\s*(\d{2}-\d{4})\s*\|\s*([\d.]+)\s*\|\s*(-?[\d.]+)\s*\|'
            matches = re.findall(pattern, html)
            say = 0
            for ay_yil, yillik_str, _ in matches:
                ay_s, yil_s = ay_yil.split("-")
                AYLIK_YILLIK[(int(yil_s), int(ay_s))] = float(yillik_str)
                say += 1
            msg = (f"✓ {say} veri TCMB'den güncellendi." if say
                   else "⚠ Veri bulunamadı.")
            fg = "#69f0ae" if say else "#ff8a65"
        except Exception as e:
            msg = f"⚠ TCMB erişim hatası: {str(e)[:50]}"
            fg = "#ff8a65"
        self.after(0, lambda: self.lbl_durum.config(text=msg, foreground=fg))

    def _temizle(self):
        self.v_tutar.set(""); self.v_musteri.set(""); self.cb_musteri.set("")
        self.tree.delete(*self.tree.get_children())
        self.lbl_erken.config(text=""); self.lbl_durum.config(text="")
        for k, a in [(self.k_bas,self.a_bas),(self.k_sim,self.a_sim),
                     (self.k_tl,self.a_tl),(self.k_pct,self.a_pct)]:
            k.config(text="—"); a.config(text="")
        self._sonuc = None
